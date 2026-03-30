import os
import glob
import xlrd
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

# ----------------------------------------------------------------------
# Вспомогательные функции
# ----------------------------------------------------------------------

def read_excel_sheet_last(file_path, sheet_name=None, has_header=False):
    """
    Читает последний лист из файла Excel (.xls или .xlsx).
    Если has_header=True, возвращает (заголовок, данные), иначе все строки.
    """
    ext = os.path.splitext(file_path)[1].lower()
    rows = []
    if ext == '.xlsx':
        wb = load_workbook(file_path, read_only=True, data_only=True)
        if sheet_name is None:
            sheet_name = wb.sheetnames[-1]
        ws = wb[sheet_name]
        for row in ws.iter_rows(values_only=True):
            rows.append(list(row))
        wb.close()
    elif ext == '.xls':
        book = xlrd.open_workbook(file_path)
        if sheet_name is None:
            sheet_name = book.sheet_names()[-1]
        sheet = book.sheet_by_name(sheet_name)
        for i in range(sheet.nrows):
            rows.append([sheet.cell_value(i, j) for j in range(sheet.ncols)])
    else:
        raise ValueError(f"Неподдерживаемый формат файла: {ext}")

    if has_header:
        if rows:
            return rows[0], rows[1:]
        else:
            return [], []
    else:
        return rows

def get_last_sheet_name(file_path):
    """
    Возвращает имя последнего листа в файле Excel.
    Поддерживает .xlsx и .xls.
    """
    ext = os.path.splitext(file_path)[1].lower()
    try:
        if ext == '.xlsx':
            wb = load_workbook(file_path, read_only=True, data_only=True)
            last_sheet = wb.sheetnames[-1]
            wb.close()
            return last_sheet
        elif ext == '.xls':
            book = xlrd.open_workbook(file_path)
            last_sheet = book.sheet_names()[-1]
            return last_sheet
        else:
            raise ValueError(f"Неподдерживаемый формат файла: {ext}")
    except Exception as e:
        print(f"Ошибка при получении имени последнего листа в файле {file_path}: {e}")
        raise

def is_last_sheet_empty(file_path, sheet_name=None):
    """
    Проверяет, пуст ли последний лист (или лист с указанным именем).
    Возвращает True, если лист не содержит ни одной непустой ячейки.
    """
    ext = os.path.splitext(file_path)[1].lower()
    try:
        if ext == '.xlsx':
            wb = load_workbook(file_path, read_only=True, data_only=True)
            if sheet_name is None:
                sheet_name = wb.sheetnames[-1]
            ws = wb[sheet_name]
            # Проверяем наличие хотя бы одной ячейки с непустым значением
            for row in ws.iter_rows(values_only=True):
                if any(cell is not None and str(cell).strip() != '' for cell in row):
                    wb.close()
                    return False
            wb.close()
            return True
        elif ext == '.xls':
            book = xlrd.open_workbook(file_path)
            if sheet_name is None:
                sheet_name = book.sheet_names()[-1]
            sheet = book.sheet_by_name(sheet_name)
            for i in range(sheet.nrows):
                for j in range(sheet.ncols):
                    val = sheet.cell_value(i, j)
                    if val is not None and str(val).strip() != '':
                        return False
            return True
        else:
            raise ValueError(f"Неподдерживаемый формат файла: {ext}")
    except Exception as e:
        print(f"Ошибка при проверке пустоты листа в файле {file_path}: {e}")
        raise

def is_row_empty(row):
    """Проверяет, является ли строка полностью пустой (все ячейки None или пустые строки)."""
    return all(cell is None or (isinstance(cell, str) and cell.strip() == '') for cell in row)

def parse_double_slash(val):
    """
    Преобразует строку вида "число//число" в кортеж (int, int) для сортировки.
    Если преобразование невозможно, возвращает (None, str(val)) для строкового сравнения.
    """
    if isinstance(val, str) and '//' in val:
        parts = val.split('//')
        if len(parts) == 2:
            try:
                first = int(parts[0].strip())
                second = int(parts[1].strip())
                return (first, second)
            except ValueError:
                # Если не числа, используем строковое сравнение
                return (None, val)
    # Для не-строк или без // используем строковое представление
    return (None, str(val))

def sort_dataframe(df, col_b, col_c):
    """
    Сортирует DataFrame по правилам:
    - сначала обычные строки (нет // и \\ в обоих столбцах) по [col_b, col_c]
    - затем строки с // или \\ в обоих столбцах по столбцу B, но с числовой сортировкой для формата "число//число"
    """
    if df.empty:
        return df

    # Проверка на наличие // или \\ в столбцах B и C
    mask_b = df[col_b].astype(str).str.contains('//', na=False) | df[col_b].astype(str).str.contains('\\\\', na=False)
    mask_c = df[col_c].astype(str).str.contains('//', na=False) | df[col_c].astype(str).str.contains('\\\\', na=False)
    mask_double = mask_b & mask_c

    # Логи: количество строк в каждой группе
    print(f"sort_dataframe: обычных строк (без // или \\\\): {len(df[~mask_double])}, "
          f"строк с // или \\\\ в обоих столбцах: {len(df[mask_double])}")

    # Логи для обратных слешей
    if df[col_b].astype(str).str.contains('\\\\', na=False).any():
        print("sort_dataframe: обнаружены строки с обратными слешами '\\' в столбце B.")
    if df[col_c].astype(str).str.contains('\\\\', na=False).any():
        print("sort_dataframe: обнаружены строки с обратными слешами '\\' в столбце C.")

    df_normal = df[~mask_double].copy()
    df_double = df[mask_double].copy()

    df_normal_sorted = df_normal.sort_values(by=[col_b, col_c])

    # Сортировка для df_double по col_b с числовым разбором
    if not df_double.empty:
        # Создаём временный столбец с ключами сортировки
        sort_keys = df_double[col_b].apply(parse_double_slash)
        df_double['__sort_key'] = sort_keys
        # Сортируем по этому ключу
        df_double_sorted = df_double.sort_values(by='__sort_key')
        # Удаляем временный столбец
        df_double_sorted = df_double_sorted.drop(columns=['__sort_key'])

        # Логирование: сколько строк распарсено успешно
        parsed_count = sum(1 for k in sort_keys if k[0] is not None)
        print(f"sort_dataframe: из {len(df_double)} строк с // или \\\\, "
              f"{parsed_count} успешно распарсены как 'число//число'.")
    else:
        df_double_sorted = df_double

    return pd.concat([df_normal_sorted, df_double_sorted], ignore_index=True)

def apply_formatting_to_sheet(ws, max_cols=5, center_cols=(2,3)):
    """
    Применяет тонкие границы ко всем строкам, у которых заполнена первая ячейка,
    в столбцах с 1 по max_cols, за исключением строк, где в третьем столбце содержится "об.".
    Для столбцов center_cols устанавливает выравнивание по центру.
    """
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_alignment = Alignment(horizontal='center')

    for row in ws.iter_rows():
        first_cell = row[0]
        # Пропускаем строки с пустой первой ячейкой
        if first_cell.value not in (None, ''):
            # Проверяем третью ячейку (индекс 2) – если это итоговая строка с "об.", границы не ставим
            third_cell = row[2] if len(row) > 2 else None
            if third_cell and third_cell.value and isinstance(third_cell.value, str) and "об." in third_cell.value:
                # Это итоговая строка – пропускаем применение границ
                continue
            # Для остальных строк применяем границы
            for cell in row[:max_cols]:
                cell.border = thin_border
                if cell.column in center_cols:
                    cell.alignment = center_alignment

def apply_formatting_to_rows(ws, start_row, data_rows, col_b_idx=1, col_c_idx=2, sample_col_idx=0):
    """
    Применяет форматирование к строкам на листе ws, начиная с start_row.
    - Серый фон для столбца A, если в столбце C = 0 или в B и C есть // или \\
    - Жирный шрифт для повторяющихся образцов (все, кроме первого вхождения)
    """
    gray_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    bold_font = Font(bold=True)
    sample_names = {}

    for i, row_data in enumerate(data_rows):
        row_num = start_row + i
        cell_c = row_data[col_c_idx] if col_c_idx < len(row_data) else None
        cell_b = row_data[col_b_idx] if col_b_idx < len(row_data) else None

        # Серый фон: условие расширено на \\
        if (cell_c == 0 or cell_c == '0') or \
           (cell_b and (('//' in str(cell_b)) or ('\\' in str(cell_b))) and
            cell_c and (('//' in str(cell_c)) or ('\\' in str(cell_c)))):
            ws.cell(row=row_num, column=1).fill = gray_fill
            # Логируем, если условие сработало из-за обратных слешей
            if cell_b and '\\' in str(cell_b) and cell_c and '\\' in str(cell_c):
                print(f"apply_formatting_to_rows: строка {row_num} получила серый фон из-за обратных слешей.")

        # Жирный шрифт для повторяющихся образцов (без изменений)
        sample = row_data[sample_col_idx] if sample_col_idx < len(row_data) else None
        if sample:
            if sample in sample_names:
                ws.cell(row=row_num, column=1).font = bold_font
            else:
                sample_names[sample] = row_num

def auto_fit_columns(worksheet, min_width=5, max_width=50, fixed_cols=None, fixed_width=5):
    """
    Автоматически подгоняет ширину всех столбцов под содержимое, кроме столбцов,
    перечисленных в fixed_cols (список букв, например ['D','E']), для которых устанавливается фиксированная ширина.
    Ширина вычисляется как максимальная длина строки в столбце (в символах),
    умноженная на коэффициент 1.2 (эмпирически подобран для openpyxl).
    Учитываются только непустые ячейки.
    Минимальная ширина задаётся параметром min_width, максимальная – max_width.
    """
    if fixed_cols is None:
        fixed_cols = []

    for col in worksheet.columns:
        if not col:
            continue
        col_letter = col[0].column_letter
        # Если столбец в списке фиксированных – устанавливаем фиксированную ширину
        if col_letter in fixed_cols:
            worksheet.column_dimensions[col_letter].width = fixed_width
            print(f"    Столбец {col_letter}: фиксированная ширина {fixed_width}")
            continue

        # Иначе – автоширина
        max_len = 0
        for cell in col:
            if cell.value is not None:
                val = str(cell.value)
                max_len = max(max_len, len(val))
        width = max_len * 1.2 if max_len > 0 else min_width
        width = max(min_width, min(max_width, width))
        worksheet.column_dimensions[col_letter].width = width
        print(f"    Столбец {col_letter}: установлена ширина {width:.1f} (макс. длина {max_len})")

# ----------------------------------------------------------------------
# Обработка основного файла
# ----------------------------------------------------------------------

def process_main_file(file_path, extra_rows=None):
    """
    Обрабатывает лист PG+ основного файла (без заголовков):
    - читает все строки как данные
    - удаляет полностью пустые строки
    - добавляет extra_rows (из специального файла PG+)
    - при необходимости расширяет число столбцов до 7
    - сортирует согласно правилам
    - перезаписывает лист с форматированием
    - добавляет итоговую строку с датой в формате ДД.ММ.ГГГГ
    - создаёт лист «Другие системы»
    Возвращает (workbook, путь_для_сохранения)
    """
    ext = os.path.splitext(file_path)[1].lower()
    TARGET_COLS = 7

    # ---- .xlsx --------------------------------------------------------
    if ext == '.xlsx':
        wb = load_workbook(file_path)
        if 'PG+' not in wb.sheetnames:
            raise ValueError("В файле отсутствует вкладка PG+.")
        ws = wb['PG+']

        # Читаем все строки как данные (первая строка тоже данные)
        data = []
        for row in ws.iter_rows(values_only=True):
            data.append(list(row))

        # Если лист пуст, создаём пустой список
        if not data:
            data = []

        # Определяем максимальное число столбцов в исходных данных
        if data:
            max_cols = max(len(row) for row in data)
        else:
            max_cols = 0

        # Приводим к TARGET_COLS (если нужно)
        if max_cols < TARGET_COLS:
            # Добавляем пустые значения справа до TARGET_COLS
            for row in data:
                row.extend([None] * (TARGET_COLS - len(row)))
            max_cols = TARGET_COLS
        elif max_cols > TARGET_COLS:
            # Обрезаем лишние столбцы
            data = [row[:TARGET_COLS] for row in data]
            max_cols = TARGET_COLS

        # Создаём искусственные заголовки (A, B, C, ...)
        headers = [chr(65 + i) for i in range(max_cols)]

        # Строим DataFrame
        df_main = pd.DataFrame(data, columns=headers)

        # Удаляем полностью пустые строки (все ячейки None или пустая строка)
        df_main = df_main.dropna(how='all')

        # Добавляем дополнительные строки из PG+, если есть
        if extra_rows:
            # Приводим extra_rows к нужной длине
            extra_rows_adj = [row[:max_cols] + [None] * (max_cols - len(row)) for row in extra_rows]
            extra_df = pd.DataFrame(extra_rows_adj, columns=headers)
            extra_df = extra_df.dropna(how='all')          # удаляем пустые и среди них
            df_main = pd.concat([df_main, extra_df], ignore_index=True)

        # Сортировка
        if max_cols > 2:
            col_b = headers[1]   # B
            col_c = headers[2]   # C
            df_sorted = sort_dataframe(df_main, col_b, col_c)
        else:
            df_sorted = df_main

        # Очищаем лист и записываем все строки (включая бывшую первую)
        ws.delete_rows(1, ws.max_row)
        for r in dataframe_to_rows(df_sorted, index=False, header=False):
            ws.append(r)

        # Применяем форматирование ко всем записанным строкам
        sorted_data = df_sorted.values.tolist()
        if sorted_data:
            apply_formatting_to_rows(ws, start_row=1, data_rows=sorted_data,
                                     col_b_idx=1, col_c_idx=2, sample_col_idx=0)

        # Подсчёт количества объектов (непустые значения в первом столбце)
        if not df_sorted.empty:
            count_objects = df_sorted.iloc[:, 0].notna().sum()
        else:
            count_objects = 0

        # Итоговая строка (дата в формате ДД.ММ.ГГГГ)
        last_data_row = ws.max_row
        summary_row = last_data_row + 1

        date_cell = ws.cell(row=summary_row, column=1, value=datetime.now().strftime("%d.%m.%Y"))
        panel_cell = ws.cell(row=summary_row, column=2, value='PG+')
        count_cell = ws.cell(row=summary_row, column=3, value=f"{count_objects} об.")

        # Выделяем жирным шрифтом
        date_cell.font = Font(bold=True)
        panel_cell.font = Font(bold=True)
        count_cell.font = Font(bold=True)
        print(f"  Итоговая строка добавлена сразу после таблицы (строка {summary_row}) без промежуточного пробела.")
        print(f"  Применён жирный шрифт к итоговой строке {summary_row} (дата, PG+, кол-во)")

        # Лист «Другие системы»
        if 'Другие системы' in wb.sheetnames:
            wb.remove(wb['Другие системы'])
        wb.create_sheet('Другие системы')

        # Устанавливаем ширину первого столбца
        ws.column_dimensions['A'].width = 12
        wb['Другие системы'].column_dimensions['A'].width = 12

        # Применяем границы и выравнивание ко всем данным на листе PG+
        apply_formatting_to_sheet(ws, max_cols=5)
        # Автоматическая подгонка ширины столбцов, фиксированная ширина для D и E
        auto_fit_columns(ws, fixed_cols=['D', 'E'])

        return wb, file_path

    # ---- .xls ---------------------------------------------------------
    elif ext == '.xls':
        # Читаем все строки (без заголовка) с листа PG+
        rows = read_excel_sheet_last(file_path, sheet_name='PG+', has_header=False)
        if not rows:
            rows = []

        # Определяем максимальное число столбцов
        if rows:
            max_cols = max(len(row) for row in rows)
        else:
            max_cols = 0

        # Приводим к TARGET_COLS
        if max_cols < TARGET_COLS:
            for row in rows:
                row.extend([None] * (TARGET_COLS - len(row)))
            max_cols = TARGET_COLS
        elif max_cols > TARGET_COLS:
            rows = [row[:TARGET_COLS] for row in rows]
            max_cols = TARGET_COLS

        # Искусственные заголовки
        headers = [chr(65 + i) for i in range(max_cols)]

        # DataFrame
        df_main = pd.DataFrame(rows, columns=headers)
        df_main = df_main.dropna(how='all')   # удаляем полностью пустые строки

        # Добавляем extra_rows, если есть
        if extra_rows:
            extra_rows_adj = [row[:max_cols] + [None] * (max_cols - len(row)) for row in extra_rows]
            extra_df = pd.DataFrame(extra_rows_adj, columns=headers)
            extra_df = extra_df.dropna(how='all')
            df_main = pd.concat([df_main, extra_df], ignore_index=True)

        # Сортировка
        if max_cols > 2:
            col_b = headers[1]
            col_c = headers[2]
            df_sorted = sort_dataframe(df_main, col_b, col_c)
        else:
            df_sorted = df_main

        # Создаём новую книгу .xlsx
        wb = Workbook()
        ws = wb.active
        ws.title = 'PG+'

        # Записываем все строки (без отдельного заголовка)
        sorted_data = df_sorted.values.tolist()
        for r_idx, row in enumerate(sorted_data, start=1):
            for c_idx, val in enumerate(row, start=1):
                ws.cell(row=r_idx, column=c_idx, value=val)

        # Форматирование
        if sorted_data:
            apply_formatting_to_rows(ws, start_row=1, data_rows=sorted_data,
                                     col_b_idx=1, col_c_idx=2, sample_col_idx=0)

        # Подсчёт объектов
        count_objects = sum(1 for row in sorted_data if row[0] not in (None, ''))

        # Итоговая строка
        summary_row = len(sorted_data) + 1
        date_cell = ws.cell(row=summary_row, column=1, value=datetime.now().strftime("%d.%m.%Y"))
        panel_cell = ws.cell(row=summary_row, column=2, value='PG+')
        count_cell = ws.cell(row=summary_row, column=3, value=f"{count_objects} об.")

        # Выделяем жирным шрифтом
        date_cell.font = Font(bold=True)
        panel_cell.font = Font(bold=True)
        count_cell.font = Font(bold=True)
        print(f"  Итоговая строка добавлена сразу после таблицы (строка {summary_row}) без промежуточного пробела.")
        print(f"  Применён жирный шрифт к итоговой строке {summary_row} (дата, PG+, кол-во) на листе PG+")

        # Лист «Другие системы»
        ws_other = wb.create_sheet('Другие системы')

        # Ширина столбцов
        ws.column_dimensions['A'].width = 12
        ws_other.column_dimensions['A'].width = 12

        # Новый путь (замена .xls на .xlsx)
        base, _ = os.path.splitext(file_path)
        output_path = base + '.xlsx'

        # Применяем границы и выравнивание ко всем данным на листе PG+
        apply_formatting_to_sheet(ws, max_cols=5)
        # Автоматическая подгонка ширины столбцов, фиксированная ширина для D и E
        auto_fit_columns(ws, fixed_cols=['D', 'E'])

        return wb, output_path

    else:
        raise ValueError("Неподдерживаемый формат файла.")



# ----------------------------------------------------------------------
# Обработка файла PG+
# ----------------------------------------------------------------------
def process_pgplus_file(file_path, rows, count_objects):
    """
    Модифицирует специальный файл PG+:
    - записывает исходные данные (5 столбцов) на последний лист
    - добавляет итоговую строку с датой, 'PG+' и количеством
    - применяет форматирование (границы, заливку, жирный шрифт)
    - переименовывает лист в текущую дату (дд.мм.гггг)
    - создаёт новый пустой лист "Новый" и делает его единственным активным
    - сохраняет файл (при необходимости конвертирует .xls в .xlsx)
    """
    ext = os.path.splitext(file_path)[1].lower()
    print(f"  Модификация специального файла PG+: {file_path}")

    if not rows:
        print("  Нет данных для записи, пропускаем модификацию.")
        return

    date_str = datetime.now().strftime("%d.%m.%Y")
    summary_row = [date_str, 'PG+', f"{count_objects} об."]

    # Определяем количество столбцов для форматирования границ (PG+ имеет 5 столбцов)
    max_cols_data = max(len(row) for row in rows) if rows else 0
    max_cols_fmt = min(max_cols_data, 5)

    if ext == '.xlsx':
        try:
            wb = load_workbook(file_path)
            last_sheet = wb.sheetnames[-1]
            ws = wb[last_sheet]

            # Очищаем лист
            ws.delete_rows(1, ws.max_row)

            # Записываем данные (5 столбцов)
            for r_idx, row in enumerate(rows, start=1):
                for c_idx, val in enumerate(row):
                    if c_idx < 5:  # защита от лишних столбцов
                        ws.cell(row=r_idx, column=c_idx+1, value=val)

            # Форматирование строк данных (используем ту же функцию, но col_b_idx=1, col_c_idx=2)
            apply_formatting_to_rows(ws, start_row=1, data_rows=rows,
                                     col_b_idx=1, col_c_idx=2, sample_col_idx=0)

            # Итоговая строка
            summary_row_num = len(rows) + 1
            for c_idx, val in enumerate(summary_row, start=1):
                cell = ws.cell(row=summary_row_num, column=c_idx, value=val)
                cell.font = Font(bold=True)
            print(f"    Применён жирный шрифт к итоговой строке {summary_row_num} (дата, PG+, кол-во)")

            # Границы
            apply_formatting_to_sheet(ws, max_cols=max_cols_fmt)
            # Автоматическая подгонка ширины столбцов, фиксированная ширина для D и E
            auto_fit_columns(ws, fixed_cols=['D', 'E'])

            # Переименовываем лист в дату (удаляем существующий с таким именем)
            if date_str in wb.sheetnames:
                print(f"    Лист '{date_str}' уже существует, удаляем.")
                wb.remove(wb[date_str])
            ws.title = date_str

            # Создаём новый пустой лист "Новый" и делаем его единственным активным
            new_sheet = "Новый"
            if new_sheet in wb.sheetnames:
                wb.remove(wb[new_sheet])
            wb.create_sheet(new_sheet)
            
            # Снимаем выделение со всех листов и выделяем только новый
            for sheet in wb.worksheets:
                sheet.sheet_view.tabSelected = False
            wb[new_sheet].sheet_view.tabSelected = True
            wb.active = wb[new_sheet]

            wb.save(file_path)
            print(f"    Файл {file_path} успешно модифицирован, активный лист: '{new_sheet}' (только он выделен).")
        except Exception as e:
            print(f"    Ошибка при модификации {file_path}: {e}")

    elif ext == '.xls':
        try:
            # Конвертируем в .xlsx
            base, _ = os.path.splitext(file_path)
            output_path = base + '.xlsx'

            wb = Workbook()
            ws = wb.active
            ws.title = "Temp"

            # Записываем данные
            for r_idx, row in enumerate(rows, start=1):
                for c_idx, val in enumerate(row):
                    if c_idx < 5:
                        ws.cell(row=r_idx, column=c_idx+1, value=val)

            apply_formatting_to_rows(ws, start_row=1, data_rows=rows,
                                     col_b_idx=1, col_c_idx=2, sample_col_idx=0)

            summary_row_num = len(rows) + 1
            for c_idx, val in enumerate(summary_row, start=1):
                cell = ws.cell(row=summary_row_num, column=c_idx, value=val)
                cell.font = Font(bold=True)
            print(f"    Применён жирный шрифт к итоговой строке {summary_row_num} (дата, PG+, кол-во)")

            apply_formatting_to_sheet(ws, max_cols=max_cols_fmt)
            auto_fit_columns(ws, fixed_cols=['D', 'E'])

            # Переименовываем лист в дату
            ws.title = date_str

            # Создаём новый пустой лист "Новый" и делаем его единственным активным
            new_sheet = "Новый"
            if new_sheet in wb.sheetnames:
                wb.remove(wb[new_sheet])
            wb.create_sheet(new_sheet)
            
            for sheet in wb.worksheets:
                sheet.sheet_view.tabSelected = False
            wb[new_sheet].sheet_view.tabSelected = True
            wb.active = wb[new_sheet]

            wb.save(output_path)
            print(f"    Файл {file_path} сконвертирован и сохранён как {output_path}, активный лист: '{new_sheet}' (только он выделен).")
        except Exception as e:
            print(f"    Ошибка при конвертации/модификации {file_path}: {e}")
    else:
        print(f"    Неподдерживаемый формат файла: {ext}")


# ----------------------------------------------------------------------
# Обработка панельных файлов
# ----------------------------------------------------------------------
def process_panel_file(file_path, sorted_data, panel_name, count_objects):
    """
    Модифицирует панельный файл:
    - записывает отсортированные данные на последний лист
    - добавляет итоговую строку с датой, панелью и количеством
    - применяет форматирование (границы, заливку, жирный шрифт)
    - переименовывает лист в текущую дату (дд.мм.гггг) (если имя не совпадает)
    - создаёт новый пустой лист "Новый" и делает его единственным активным
    - сохраняет файл (при необходимости конвертирует .xls в .xlsx)
    """
    ext = os.path.splitext(file_path)[1].lower()
    print(f"  Модификация панельного файла: {file_path}")
    print(f"  Получено {len(sorted_data)} строк данных для записи.")

    if not sorted_data:
        print("  Нет данных для записи, пропускаем модификацию.")
        return

    date_str = datetime.now().strftime("%d.%m.%Y")
    summary_row = [date_str, panel_name, f"{count_objects} об."]

    # Определяем количество столбцов для форматирования границ (не более 5)
    max_cols_data = max(len(row) for row in sorted_data) if sorted_data else 0
    max_cols_fmt = min(max_cols_data, 5)

    if ext == '.xlsx':
        try:
            wb = load_workbook(file_path)
            last_sheet = wb.sheetnames[-1]
            ws = wb[last_sheet]
            print(f"    Последний лист перед обработкой: '{last_sheet}'")

            # Очищаем лист
            ws.delete_rows(1, ws.max_row)
            print(f"    Лист очищен, текущее количество строк: {ws.max_row}")

            # Записываем данные
            for r_idx, row in enumerate(sorted_data, start=1):
                for c_idx, val in enumerate(row, start=1):
                    ws.cell(row=r_idx, column=c_idx, value=val)
            print(f"    Записано {len(sorted_data)} строк данных.")

            # Форматирование строк данных
            apply_formatting_to_rows(ws, start_row=1, data_rows=sorted_data,
                                     col_b_idx=1, col_c_idx=2, sample_col_idx=0)

            # Итоговая строка
            summary_row_num = len(sorted_data) + 1
            for c_idx, val in enumerate(summary_row, start=1):
                cell = ws.cell(row=summary_row_num, column=c_idx, value=val)
                cell.font = Font(bold=True)
            print(f"    Применён жирный шрифт к итоговой строке {summary_row_num} (дата, панель, кол-во)")

            # Границы и автоширина
            apply_formatting_to_sheet(ws, max_cols=max_cols_fmt)
            auto_fit_columns(ws, fixed_cols=['D', 'E'])

            # Переименовываем лист в дату, только если имя не совпадает
            if ws.title != date_str:
                if date_str in wb.sheetnames:
                    print(f"    Лист '{date_str}' уже существует, удаляем.")
                    wb.remove(wb[date_str])
                ws.title = date_str
                print(f"    Лист переименован в '{date_str}'")
            else:
                print(f"    Лист уже имеет имя '{date_str}', переименование не требуется.")

            # Создаём новый пустой лист "Новый" и делаем его единственным активным
            new_sheet = "Новый"
            if new_sheet in wb.sheetnames:
                wb.remove(wb[new_sheet])
            wb.create_sheet(new_sheet)
            
            # Снимаем выделение со всех листов и выделяем только новый
            for sheet in wb.worksheets:
                sheet.sheet_view.tabSelected = False
            wb[new_sheet].sheet_view.tabSelected = True
            wb.active = wb[new_sheet]

            wb.save(file_path)
            print(f"    Файл {file_path} успешно модифицирован, активный лист: '{new_sheet}'")
            print(f"    Теперь в файле листы: {wb.sheetnames}")
        except Exception as e:
            print(f"    Ошибка при модификации {file_path}: {e}")

    elif ext == '.xls':
        try:
            # Конвертируем в .xlsx
            base, _ = os.path.splitext(file_path)
            output_path = base + '.xlsx'

            wb = Workbook()
            ws = wb.active
            ws.title = "Temp"
            print(f"    Создана новая книга для конвертации: {output_path}")

            # Записываем данные
            for r_idx, row in enumerate(sorted_data, start=1):
                for c_idx, val in enumerate(row, start=1):
                    ws.cell(row=r_idx, column=c_idx, value=val)
            print(f"    Записано {len(sorted_data)} строк данных.")

            apply_formatting_to_rows(ws, start_row=1, data_rows=sorted_data,
                                     col_b_idx=1, col_c_idx=2, sample_col_idx=0)

            summary_row_num = len(sorted_data) + 1
            for c_idx, val in enumerate(summary_row, start=1):
                cell = ws.cell(row=summary_row_num, column=c_idx, value=val)
                cell.font = Font(bold=True)
            print(f"    Применён жирный шрифт к итоговой строке {summary_row_num} (дата, панель, кол-во)")

            apply_formatting_to_sheet(ws, max_cols=max_cols_fmt)
            auto_fit_columns(ws, fixed_cols=['D', 'E'])

            # Переименовываем лист в дату (в новой книге лист "Temp", конфликта нет)
            ws.title = date_str
            print(f"    Лист переименован в '{date_str}'")

            # Создаём новый пустой лист "Новый" и делаем его единственным активным
            new_sheet = "Новый"
            if new_sheet in wb.sheetnames:
                wb.remove(wb[new_sheet])
            wb.create_sheet(new_sheet)
            
            for sheet in wb.worksheets:
                sheet.sheet_view.tabSelected = False
            wb[new_sheet].sheet_view.tabSelected = True
            wb.active = wb[new_sheet]

            wb.save(output_path)
            print(f"    Файл {file_path} сконвертирован и сохранён как {output_path}, активный лист: '{new_sheet}'")
            print(f"    Теперь в файле листы: {wb.sheetnames}")
        except Exception as e:
            print(f"    Ошибка при конвертации/модификации {file_path}: {e}")
    else:
        print(f"    Неподдерживаемый формат файла: {ext}")

# ----------------------------------------------------------------------
# Основная программа
# ----------------------------------------------------------------------
def main():
    # 1. Самый новый файл в текущей папке
    files = glob.glob('*.xls') + glob.glob('*.xlsx')
    if not files:
        print("Не найдено файлов Excel в текущей папке.")
        return
    latest_file = max(files, key=os.path.getctime)
    print(f"Обрабатывается файл: {latest_file}")

    # 2. Путь к папке с панельными файлами
    panel_dir = r"U:\ГЕНЕТИКА НОВОЕ ЗДАНИЕ\АНАЛИЗАТОР\Работа\2-Реал-тайм\2026"

    # 3. Специальный файл PG+ (берём первый найденный)
    pg_plus_files = glob.glob(os.path.join(panel_dir, "*PG+*.xlsx")) + \
                    glob.glob(os.path.join(panel_dir, "*PG+*.xls"))
    extra_rows = None
    pg_file = None
    if pg_plus_files:
        pg_file = pg_plus_files[0]
        print(f"Найден специальный файл PG+: {pg_file}")
        rows = read_excel_sheet_last(pg_file, has_header=False)
        if rows:
            # Удаляем полностью пустые строки
            rows = [row for row in rows if not is_row_empty(row)]
            if rows:
                # Подсчитываем количество объектов (непустые в первом столбце)
                count_objects_pg = sum(1 for row in rows if row[0] not in (None, ''))
                # Модифицируем сам файл PG+ (добавляем итоговую строку, новый лист)
                process_pgplus_file(pg_file, rows, count_objects_pg)

                # Преобразуем строки в формат с 7 столбцами для добавления в основной файл
                extra_rows = []
                for row in rows:
                    new_row = [None] * 7
                    if len(row) > 0: new_row[0] = row[0]  # Sample Name
                    if len(row) > 1: new_row[1] = row[1]  # DNA to D1
                    if len(row) > 2: new_row[2] = row[2]  # Diluent to D1
                    if len(row) > 3: new_row[5] = row[3]  # эксперт -> F
                    if len(row) > 4: new_row[6] = row[4]  # примечания -> G
                    extra_rows.append(new_row)
            else:
                print("Внимание: файл PG+ содержит только пустые строки, не обрабатывается.")
        else:
            print("Внимание: файл PG+ пуст.")

    # 4. Обработка основного файла с учётом extra_rows
    # Добавляем обработку ошибок, чтобы при отсутствии вкладки PG+ программа не закрывалась сразу
    try:
        wb, output_path = process_main_file(latest_file, extra_rows)
        ws_other = wb['Другие системы']
    except Exception as e:
        print(f"\nОШИБКА при обработке основного файла: {e}")
        # Проверяем, связана ли ошибка с отсутствием вкладки PG+
        if "PG+" in str(e):
            print("В основном файле отсутствует вкладка 'PG+'. Это критическая ошибка.")
        else:
            print("Произошла ошибка при обработке основного файла.")
        print("\nНажмите Enter для выхода...")
        input()          # Ожидание нажатия любой клавиши (Enter)
        return           # Завершаем программу, так как дальнейшая работа невозможна

    # 5. Обработка всех панельных файлов (кроме PG+)
    all_panel_files = glob.glob(os.path.join(panel_dir, "На *.xlsx")) + \
                      glob.glob(os.path.join(panel_dir, "На *.xls"))
    panel_files = [f for f in all_panel_files if "PG+" not in os.path.basename(f)]

    # Множество файлов, для которых уже была выполнена проверка имени листа
    # (повторная обработка после предупреждения)
    retry_checked = set()

    for panel_file in panel_files:
        # Флаг: пропустить проверку имени листа для этого файла
        skip_name_check = panel_file in retry_checked

        while True:
            try:
                # 1. Проверка имени последнего листа (только если не пропущена)
                if not skip_name_check:
                    last_sheet_name = get_last_sheet_name(panel_file)
                    is_empty = is_last_sheet_empty(panel_file, last_sheet_name)
                    print(f"Файл: {os.path.basename(panel_file)}, последний лист: '{last_sheet_name}', пуст: {is_empty}")

                    # Если имя нестандартное (не входит в список)
                    if last_sheet_name not in ["Новый", "New", "Лист1", "Sheet1"]:
                        print(f"ВНИМАНИЕ: В файле {os.path.basename(panel_file)} последний лист переименован!")
                        input("Проверьте файл, после чего нажмите Enter для продолжения...")
                        # Запоминаем, что для этого файла проверку больше не делаем
                        retry_checked.add(panel_file)
                        # Повторяем обработку этого же файла с начала, пропуская проверку имени
                        # Для этого выходим из текущей итерации while, но остаёмся в цикле while
                        # и устанавливаем skip_name_check = True
                        skip_name_check = True
                        continue   # переходим к следующей итерации while

                    # Имя стандартное, но лист пустой – пропускаем файл
                    if is_empty:
                        print(f"Последний лист '{last_sheet_name}' в файле {os.path.basename(panel_file)} пустой. Пропускаем файл.")
                        break   # выходим из while, переходим к следующему файлу

                # 2. Если дошли сюда – либо проверка пропущена, либо имя стандартное и лист не пустой
                # Читаем данные с последнего листа (используем существующую функцию)
                rows = read_excel_sheet_last(panel_file, has_header=False)
                if not rows:
                    print(f"  Файл {os.path.basename(panel_file)} не содержит данных (пустой лист).")
                    break

                # Удаляем полностью пустые строки
                rows = [row for row in rows if not is_row_empty(row)]

                if not rows:
                    print(f"  Файл {os.path.basename(panel_file)} не содержит значимых данных (все строки пусты).")
                    break

                print(f"  Найдено записей после очистки: {len(rows)}")

                # 3. Проверка количества записей (1-2) – оставляем как было
                if len(rows) in (1, 2):
                    word = "запись" if len(rows) == 1 else "записи"
                    resp = input(f"В файле {os.path.basename(panel_file)} {len(rows)} {word}. Обрабатывать (д/н)? ")
                    print(f"Пользователь ввёл: '{resp}'")
                    if resp.lower() != 'д':
                        print(f"Пропускаем файл {os.path.basename(panel_file)} по решению пользователя.")
                        break
                    else:
                        print(f"Продолжаем обработку файла {os.path.basename(panel_file)}.")

                # 4. Обработка данных (приведение к единой длине, сортировка, запись на лист "Другие системы")
                max_cols = max(len(row) for row in rows)
                data = [row + [None] * (max_cols - len(row)) for row in rows]
                df = pd.DataFrame(data)

                if df.shape[1] > 2:
                    df_sorted = sort_dataframe(df, 1, 2)
                else:
                    df_sorted = df

                sorted_data = df_sorted.values.tolist()

                # Подсчёт количества объектов
                count_objects = sum(1 for row in sorted_data if row[0] not in (None, ''))

                # Имя панели
                base_name = os.path.basename(panel_file)
                if '-' in base_name:
                    panel_name = base_name.split('-')[0]
                else:
                    panel_name = base_name.replace('.xlsx', '').replace('.xls', '')

                # Итоговая строка
                summary = [datetime.now().strftime("%d.%m.%Y"), panel_name, f"{count_objects} об."]

                # Запись на лист "Другие системы"
                start_row = ws_other.max_row + 1
                print(f"  Начальная строка для блока: {start_row}")

                for r_idx, row_data in enumerate(sorted_data):
                    row_num = start_row + r_idx
                    for c_idx, val in enumerate(row_data):
                        ws_other.cell(row=row_num, column=c_idx+1, value=val)

                # Форматирование внутри блока
                apply_formatting_to_rows(ws_other, start_row, sorted_data,
                                         col_b_idx=1, col_c_idx=2, sample_col_idx=0)

                # Итоговая строка
                summary_row = start_row + len(sorted_data)
                for c_idx, val in enumerate(summary):
                    cell = ws_other.cell(row=summary_row, column=c_idx+1, value=val)
                    cell.font = Font(bold=True)
                print(f"  Применён жирный шрифт к итоговой строке {summary_row} (дата, панель, кол-во) на листе 'Другие системы'")

                # Пустая строка после блока
                ws_other.append([None] * 5)
                print(f"  Добавлена пустая строка после блока. Теперь max_row = {ws_other.max_row}")

                # 5. Модификация самого панельного файла
                try:
                    print(f"  Вызов process_panel_file для {os.path.basename(panel_file)} с {len(sorted_data)} строками")
                    process_panel_file(panel_file, sorted_data, panel_name, count_objects)
                    print(f"  Файл {os.path.basename(panel_file)} успешно обработан.")
                except Exception as e:
                    print(f"  Ошибка при модификации панельного файла {panel_file}: {e}")

                # Успешно обработали – выходим из цикла while
                break

            except Exception as e:
                print(f"Критическая ошибка при обработке {panel_file}: {e}")
                # При ошибке прерываем цикл для этого файла
                break

    # Удаление пустой строки в начале листа (если есть)
    if ws_other.max_row >= 1:
        # Проверяем, пустая ли первая строка (все ячейки в первых трёх столбцах)
        first_row_values = [ws_other.cell(row=1, column=col).value for col in range(1, 4)]
        if all(v is None or (isinstance(v, str) and v.strip() == '') for v in first_row_values):
            print("Обнаружена пустая строка в начале листа 'Другие системы'. Удаляем её.")
            ws_other.delete_rows(1)
            print(f"После удаления первой строки max_row = {ws_other.max_row}")

    # 6. Сохранение финальной книги
    # Применяем форматирование границ к листу "Другие системы"
    apply_formatting_to_sheet(ws_other, max_cols=3)
    # Автоматическая подгонка ширины столбцов, фиксированная ширина для D и E
    # (если на листе нет D/E, они проигнорируются)
    auto_fit_columns(ws_other, fixed_cols=['D', 'E'])

    # Сохраняем
    wb.save(output_path)
    print(f"Финальный файл сохранён: {output_path}")
    print("Обработка завершена.")

if __name__ == "__main__":
    main()